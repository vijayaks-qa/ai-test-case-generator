from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

def save_to_excel(test_cases, feature):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test cases"

    #define border
    thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
    )
    
    # Header
    ws.append([
        "Test Case ID",
        "Title",
        "Scenario",
        "Preconditions",
        "Test Steps",
        "Expected Result",
        "Priority",
        "Test Type"
    ])
    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    # Rows
    for tc in test_cases:
        ws.append([
            tc["Test Case ID"],
            tc["Title"],
            tc["Scenario"],
            tc["Preconditions"],
            "\n".join(f"{i + 1}. {step}"
    for i, step in enumerate(tc["Test Steps"])),
            tc["Expected Result"],
            tc["Priority"],
            tc["Test Type"]
        ])

    output_folder = "output"
    os.makedirs(output_folder, exist_ok=True)

    filename = os.path.join(
        output_folder,
        f"{feature.replace(' ', '_')}_TestCases.xlsx"
    )

    #apply border
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = max_length + 5

    # Wrap text and align to top
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

    # Adjust row height
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 70

    wb.save(filename)
    print(f"✅ Excel file '{filename}' created successfully.")
    return filename